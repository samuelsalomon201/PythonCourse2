import openpyxl

workbook = openpyxl.load_workbook("E:\Employees.xlsx")
print(workbook.properties)
print(workbook.sheetnames)
print(workbook.active)
sheet = workbook['EmployeeData']
workbook.create_sheet('TestSheet')
workbook.save("EmployeeData")

sheet = workbook['TestSheet']
workbook.remove(sheet)

workbook.save("EmployeeData")

sheet = workbook['EmployeeData']

print(sheet.title)

print(dir(sheet))


print(sheet.active_cell)

print(sheet.dimensions)

print(sheet.max_row)

print(sheet.max_column)

print(sheet.sheet_properties)

print(sheet.sheet_view)


for i in sheet.values:
    print(i)