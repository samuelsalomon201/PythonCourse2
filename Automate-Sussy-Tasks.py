import openpyxl

workbook = openpyxl.load_workbook("E:\samuel\Employees.xlsx")
print(workbook.properties)
print(workbook.sheetnames)
print(workbook.active)
sheet = workbook['EmployeeData']
workbook.create_sheet('Amogus Test')
# This Give Big Error
# workbook.save("E:\samuel\Employees.xlsx")
# sheet = workbook['TestSheet']
# workbook.remove(sheet)
# del workbook['Amogus Test']
# workbook.save('E:\samuel\Employees.xlsx')
sheet = workbook['EmployeeData']
print(sheet.title)
print(dir(sheet))
print(sheet.active_cell)
print(sheet.dimensions)
print(sheet.sheet_view)
